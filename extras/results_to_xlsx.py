import os
import argparse
import yaml
import numpy as np

import openpyxl
from openpyxl import Workbook

from common.utils import load_meta

if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument('--folder', default='results', type=str)
    parser.add_argument('--del_empty_dir', action='store_true')

    args = parser.parse_args()

    metas = {}

    for subdir, dirs, files in os.walk(args.folder):
        if len(dirs):
            continue

        if not len(files):
            if args.del_empty_dir:
                print('deleting folder %s' % subdir)
                os.rmdir(os.path.abspath(subdir))

        if 'model.h5' not in files:
            print('model.h5 not found in %s' % subdir)
            continue

        meta = load_meta(os.path.join(subdir, 'model.h5'))
        metas[subdir.split(os.sep)[-1]] = meta

    training_args = list(set([arg for model in metas for arg in
                             metas[model]['training_args']]))

    datasets = {}
    for model in metas:
        args = metas[model]['training_args']
        meta = metas[model]

        key = None
        for search_key in ('dataset', 'train'):
            if (search_key in args and args[search_key] is not None):
                key = args[search_key].split(os.sep)[-2]
                break

        key = key or 'unknown'

        if key not in datasets:
            datasets[key] = {}

        datasets[key][model] = meta

    wb = Workbook()

columns = ['path'] + ['epoch', 'best_val_ler'] + training_args

for name in datasets:
    ws = wb.create_sheet(name)

    cell_range = ws['A1':'%s1' % openpyxl.utils.get_column_letter(len(columns))][0]

    for i, cell in zip(range(len(cell_range)), cell_range):
        cell.value = columns[i]

    for row, (model, meta) in enumerate(datasets[name].items(), start=2):

        ws['A%d' % row] = model
        for key in ('epoch', 'epochs'):
            if key in meta:
                ws['B%d' % row] = meta[key][np.argmin(meta['val_decoder_ler'])]
                break
        ws['C%d' % row] = np.min(meta['val_decoder_ler'])

        for arg, val in meta['training_args'].items():
            col = openpyxl.utils.get_column_letter(training_args.index(arg) + 4)

            ws['%s%d' % (col, row)] = val


wb.save('results.xlsx')
